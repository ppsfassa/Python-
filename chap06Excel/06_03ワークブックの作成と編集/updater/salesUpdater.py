import openpyxl

book = openpyxl.load_workbook('惣菜売上.xlsx')   # Workbookオブジェクトを生成
sheet = book.get_sheet_by_name('Sheet1')         # Worksheetオブジェクトを生成

PRICE_UPDATES = {'秋刀魚の竜田揚げ': 3.66,       # 変更する商品名と変更後の単価を辞書に登録
                 '鶏つみれ大根': 2.78,
                 'きのこの白和え': 2.16}

# 該当する商品の単価を更新する
for row_num in range(2, sheet.max_row + 1):      # 先頭行を除いて2行目からループを開始
    name = sheet.cell(row=row_num,               # 行番号を指定
                      column=1                   # 商品名が登録されている1列目を指定
                      ).value                    # 商品名を取得する
    if name in PRICE_UPDATES:                    # 商品名がPRICE_UPDATESの商品名と一致するか調べる
                                                 # 商品名が一致したらPRICE_UPDATESの単価に更新する
        sheet.cell(row=row_num,                  # 行番号を指定
                   column=2                      # 単価が登録されている2行目を指定
                   ).value = PRICE_UPDATES[name] # PRICE_UPDATESのnameキーの値に更新

book.save('惣菜売上_updated.xlsx')               # 更新したワークブックを別名で保存

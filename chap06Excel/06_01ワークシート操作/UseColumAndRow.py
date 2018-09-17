import openpyxl

book = openpyxl.load_workbook('地形別面積.xlsx') # Excelブックを取得
sheet = book.get_sheet_by_name('Sheet1')         # Sheet1を取得

print(sheet.cell(row=3,       # 行を指定
                 column=1     # 列を指定                 
                 ).value)

for i in range(2, 8):         # 2行目から7行目までを繰り返す
    print(i,                  # 行番号
          sheet.cell(row=i,   # 2～7が順番に代入される
                     column=1 # 列は2で固定
                     ).value)


for i in range(2, 8, 2):      # 2行目から7行目までを1行おきに繰り返す
    print(i,                  # 行番号
          sheet.cell(row=i,   # 2～7まで1つおきに代入される
                     column=1 # 列は2で固定
                     ).value)

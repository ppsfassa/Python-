import openpyxl
# get_column_letter、column_index_from_stringをインポート
from openpyxl.utils import get_column_letter, column_index_from_string

print('列の文字', get_column_letter(1))           # 列番号から列の文字を取得
print('列番号', column_index_from_string('A'))    # 列の文字から列番号を取得     

book  = openpyxl.load_workbook('地形別面積.xlsx') # Excelブックを取得
sheet = book.get_sheet_by_name('Sheet1')          # Sheet1を取得
print('最終列の列文字->',
      get_column_letter(sheet.max_column))        # 最終列の列文字を取得

import openpyxl

book  = openpyxl.load_workbook('地形別面積.xlsx') # Excelブックを取得
sheet = book.get_sheet_by_name('Sheet1')         # Sheet1を取得

print('最大行数->',
      sheet.max_row,   # 表の行数を取得
      '最大列数->',
      sheet.max_column # 表の列数を取得
      )
